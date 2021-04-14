# -*- coding:utf-8 -*-



# 打开Excel文件读取数据

from openpyxl import load_workbook,Workbook

'''
想要学习Python？Python学习交流群：984632579满足你的需求，资料都已经上传群文件，可以自行下载！
'''
file_path = r'C:\Users\liuyechun\Desktop\工作爬取\2021-1-19-吉林智联.xlsx'

# 打开一个workbook
wb = load_workbook(file_path)

# 获取当前活跃的worksheet,默认就是第一个worksheet
# ws = wb.active

# 当然也可以使用下面的方法

# 获取所有表格(worksheet)的名字
sheets = wb.get_sheet_names()
# 第一个表格的名称
sheet_first = sheets[0]
# 获取特定的worksheet
ws = wb.get_sheet_by_name(sheet_first)

# 获取表格所有行和列，两者都是可迭代的
rows = ws.rows
columns = ws.columns

gongsi_list = []
# 迭代所有的行
for row in rows:
    line = [col.value for col in row]
    gongsi = line[-1]
    if gongsi == '公司':
        continue
    gongsi_list.append(gongsi)
print(gongsi_list)


"""
企查查  联系方式
"""

import requests
import time
import parsel

# gongsi_list = [ '沈阳运通宸宝汽车销售服务有限公司', '沈阳华宝汽车销售服务有限公司', '沈阳宇宸汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '北京英跑汽车销售服务有限公司', '沈阳沈东金廊华林汽车销售服务有限公司', '朝阳中升正茂汽车销售服务有限公司', '辽宁万亿达汽车销售服务有限公司', '大连中升宏达汽车销售服务有限公司', '大连中升汇驰汽车销售服务有限公司', '大连中升汇驰汽车销售服务有限公司', '铁岭和成汽车销售服务有限公司', '沈阳华宝汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '沈阳中晨东怡汽车销售服务有限公司', '沈阳金廊雷克萨斯汽车销售服务有限公司', '沈阳南光丰田汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '大连新美联盛金汽车销售服务有限公司', '辽宁广鑫隆汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '沈阳金廊华林汽车销售服务有限公司', '沈阳中资华大汽车销售服务有限公司', '大连新美联盛金汽车销售服务有限公司', '辽宁新世纪汽车销售服务有限公司', '辽宁顺亿达汽车销售服务有限公司', '爱威马(辽宁)汽车销售服务有限责任公司', '大连中升盛通汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '朝阳四隆之星汽车销售服务有限公司', '大连保税区华通丰田汽车销售服务有限公司', '沈阳中晨先锋丰田汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '大连东升汽车销售服务有限公司', '大连保税区华通丰田汽车销售服务有限公司', '大连保税区华通丰田汽车销售服务有限公司', '大连中升丰田汽车销售服务有限公司', '大连中升之宝汽车销售服务有限公司（BMW宝马店）', '沈阳中晨诚隆汽车销售服务有限公司', '沈阳中晨先锋丰田汽车销售服务有限公司', '沈阳中晨诚隆汽车销售服务有限公司', '辽宁上辽汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '大连中升日产汽车销售服务有限公司', '沈阳南星行汽车销售服务有限公司', '丹东江通汽车销售服务有限公司', '沈阳金廊华林汽车销售服务有限公司', '抚顺上辽汽车销售服务有限公司', '沈阳中资华大汽车销售服务有限公司', '大连东升汽车销售服务有限公司', '沈阳中资华大汽车销售服务有限公司', '铁岭和成汽车销售服务有限公司', '辽宁鑫盛特汽车销售服务有限公司', '大连福友达汽车销售服务有限公司', '沈阳华宝汽车销售服务有限公司', '大连中升汇驰汽车销售服务有限公司', '营口大通汽车销售服务有限公司', '辽宁大地汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '辽宁上辽汽车销售服务有限公司', '爱威马(辽宁)汽车销售服务有限责任公司', '本溪创为汽车销售服务有限公司', '沈阳金廊雷克萨斯汽车销售服务有限公司', '辽宁盛世福丰汽车销售服务有限公司', '沈阳华宝汽车销售服务有限公司（BMW宝马店）', '锦州市欧亚汽车销售服务有限责任公司', '辽宁鑫汇通汽车销售服务有限公司', '营口君恒汽车销售服务有限公司', '沈阳运通宸宝汽车销售服务有限公司', '葫芦岛华宝汽车销售服务有限公司（BMW宝马店）', '辽阳英迪汽车销售服务有限公司', '营口君恒汽车销售服务有限公司', '鞍山和卓汽车销售服务有限公司', '鞍山和卓汽车销售服务有限公司', '大连亿霖汽车销售服务有限公司', '大连嘉和润诚汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳市宝晋汽车销售服务有限公司（BMW宝马店）', '辽宁路赛得安特汽车销售服务有限公司', '沈阳中晨诚隆汽车销售服务有限公司', '沈阳中晨东怡汽车销售服务有限公司', '大连中升宏达汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '大连骏德汽车销售服务有限公司', '大连中升之宝汽车销售服务有限公司（BMW宝马店）', '大连中升汽车销售服务有限公司', '爱威马(辽宁)汽车销售服务有限责任公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳喜悦汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '沈阳路丰源汽车销售服务有限公司', '沈阳宇宸汽车销售服务有限公司', '辽宁奥通汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '鞍山和卓汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳南光丰田汽车销售服务有限公司', '大连金廊华林汽车销售服务有限公司', '鞍山四隆之星汽车销售服务有限公司', '辽宁新世纪汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '沈阳中升仕豪汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '大连中升之宝汽车销售服务有限公司（BMW宝马店）', '辽宁奥通汽车销售服务有限公司', '大连裕迪汽车销售服务有限公司', '沈阳喜悦汽车销售服务有限公司', '辽宁顺亿达汽车销售服务有限公司', '辽宁奥通汽车销售服务有限公司', '大连中升汽车销售服务有限公司', '大连壁虎汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '大连华菱盛大汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '辽宁众兴汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '大连万海汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '大连汇华丰田汽车销售服务有限公司', '大连领航汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '铁岭天广吉祥汽车销售服务有限公司', '辽宁利丰星行汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '辽宁神州红旗汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '沈阳金廊雷克萨斯汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '沈阳中晨东怡汽车销售服务有限公司', '沈阳金廊雷克萨斯汽车销售服务有限公司', '沈阳南光丰田汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '大连新美联盛金汽车销售服务有限公司', '辽宁广鑫隆汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '沈阳金廊华林汽车销售服务有限公司', '沈阳中资华大汽车销售服务有限公司', '大连新美联盛金汽车销售服务有限公司', '辽宁新世纪汽车销售服务有限公司', '辽宁顺亿达汽车销售服务有限公司', '爱威马(辽宁)汽车销售服务有限责任公司', '大连中升盛通汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '朝阳四隆之星汽车销售服务有限公司', '大连保税区华通丰田汽车销售服务有限公司', '沈阳中晨先锋丰田汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '大连东升汽车销售服务有限公司', '大连保税区华通丰田汽车销售服务有限公司', '大连保税区华通丰田汽车销售服务有限公司', '大连中升丰田汽车销售服务有限公司', '大连中升之宝汽车销售服务有限公司（BMW宝马店）', '沈阳中晨诚隆汽车销售服务有限公司', '沈阳中晨先锋丰田汽车销售服务有限公司', '沈阳中晨诚隆汽车销售服务有限公司', '辽宁上辽汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '大连中升日产汽车销售服务有限公司', '沈阳南星行汽车销售服务有限公司', '丹东江通汽车销售服务有限公司', '沈阳金廊华林汽车销售服务有限公司', '抚顺上辽汽车销售服务有限公司', '沈阳中资华大汽车销售服务有限公司', '大连东升汽车销售服务有限公司', '沈阳中资华大汽车销售服务有限公司', '铁岭和成汽车销售服务有限公司', '辽宁鑫盛特汽车销售服务有限公司', '大连福友达汽车销售服务有限公司', '沈阳华宝汽车销售服务有限公司', '大连中升汇驰汽车销售服务有限公司', '营口大通汽车销售服务有限公司', '辽宁大地汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '辽宁上辽汽车销售服务有限公司', '爱威马(辽宁)汽车销售服务有限责任公司', '本溪创为汽车销售服务有限公司', '沈阳金廊雷克萨斯汽车销售服务有限公司', '辽宁盛世福丰汽车销售服务有限公司', '沈阳华宝汽车销售服务有限公司（BMW宝马店）', '锦州市欧亚汽车销售服务有限责任公司', '辽宁鑫汇通汽车销售服务有限公司', '营口君恒汽车销售服务有限公司', '沈阳运通宸宝汽车销售服务有限公司', '葫芦岛华宝汽车销售服务有限公司（BMW宝马店）', '辽阳英迪汽车销售服务有限公司', '营口君恒汽车销售服务有限公司', '鞍山和卓汽车销售服务有限公司', '大连亿霖汽车销售服务有限公司', '大连嘉和润诚汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳市宝晋汽车销售服务有限公司（BMW宝马店）', '辽宁路赛得安特汽车销售服务有限公司', '沈阳中晨诚隆汽车销售服务有限公司', '沈阳中晨东怡汽车销售服务有限公司', '大连中升宏达汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '大连骏德汽车销售服务有限公司', '大连中升之宝汽车销售服务有限公司（BMW宝马店）', '大连中升汽车销售服务有限公司', '爱威马(辽宁)汽车销售服务有限责任公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳喜悦汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '沈阳路丰源汽车销售服务有限公司', '沈阳宇宸汽车销售服务有限公司', '辽宁奥通汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '鞍山和卓汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳南光丰田汽车销售服务有限公司', '大连金廊华林汽车销售服务有限公司', '鞍山四隆之星汽车销售服务有限公司', '辽宁新世纪汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '沈阳中升仕豪汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '大连中升之宝汽车销售服务有限公司（BMW宝马店）', '辽宁奥通汽车销售服务有限公司', '大连裕迪汽车销售服务有限公司', '沈阳喜悦汽车销售服务有限公司', '辽宁顺亿达汽车销售服务有限公司', '辽宁奥通汽车销售服务有限公司', '大连中升汽车销售服务有限公司', '大连壁虎汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '大连华菱盛大汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '辽宁众兴汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '大连万海汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '大连汇华丰田汽车销售服务有限公司', '大连领航汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '铁岭天广吉祥汽车销售服务有限公司', '辽宁利丰星行汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '辽宁神州红旗汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '沈阳金廊雷克萨斯汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '沈阳金廊华林汽车销售服务有限公司', '沈阳中资华大汽车销售服务有限公司', '大连新美联盛金汽车销售服务有限公司', '辽宁新世纪汽车销售服务有限公司', '辽宁顺亿达汽车销售服务有限公司', '爱威马(辽宁)汽车销售服务有限责任公司', '大连中升盛通汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '朝阳四隆之星汽车销售服务有限公司', '大连保税区华通丰田汽车销售服务有限公司', '沈阳中晨先锋丰田汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '大连东升汽车销售服务有限公司', '大连保税区华通丰田汽车销售服务有限公司', '大连保税区华通丰田汽车销售服务有限公司', '大连中升丰田汽车销售服务有限公司', '沈阳中晨诚隆汽车销售服务有限公司', '沈阳中晨先锋丰田汽车销售服务有限公司', '沈阳中晨诚隆汽车销售服务有限公司', '辽宁上辽汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '大连中升日产汽车销售服务有限公司', '沈阳南星行汽车销售服务有限公司', '丹东江通汽车销售服务有限公司', '沈阳金廊华林汽车销售服务有限公司', '抚顺上辽汽车销售服务有限公司', '沈阳中资华大汽车销售服务有限公司', '大连东升汽车销售服务有限公司', '沈阳中资华大汽车销售服务有限公司', '铁岭和成汽车销售服务有限公司', '辽宁鑫盛特汽车销售服务有限公司', '大连福友达汽车销售服务有限公司', '沈阳华宝汽车销售服务有限公司', '大连中升汇驰汽车销售服务有限公司', '营口大通汽车销售服务有限公司', '辽宁大地汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '辽宁上辽汽车销售服务有限公司', '爱威马(辽宁)汽车销售服务有限责任公司', '本溪创为汽车销售服务有限公司', '大连中升汽车销售服务有限公司', '爱威马(辽宁)汽车销售服务有限责任公司', '辽宁路赛得安特汽车销售服务有限公司', '大连亿霖汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '沈阳路丰源汽车销售服务有限公司', '沈阳宇宸汽车销售服务有限公司', '辽宁奥通汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '鞍山和卓汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳南光丰田汽车销售服务有限公司', '鞍山四隆之星汽车销售服务有限公司', '大连骏德汽车销售服务有限公司', '辽宁新世纪汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '沈阳中升仕豪汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '沈阳喜悦汽车销售服务有限公司', '大连中升之宝汽车销售服务有限公司（BMW宝马店）', '辽宁奥通汽车销售服务有限公司', '大连金廊华林汽车销售服务有限公司', '大连裕迪汽车销售服务有限公司', '沈阳喜悦汽车销售服务有限公司', '辽宁顺亿达汽车销售服务有限公司', '辽宁奥通汽车销售服务有限公司', '大连壁虎汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '大连华菱盛大汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '辽宁众兴汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '大连中升汽车销售服务有限公司', '大连万海汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '大连汇华丰田汽车销售服务有限公司', '大连领航汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '铁岭天广吉祥汽车销售服务有限公司', '辽宁利丰星行汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '辽宁神州红旗汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '沈阳金廊雷克萨斯汽车销售服务有限公司', '大连东升汽车销售服务有限公司', '大连保税区华通丰田汽车销售服务有限公司', '大连保税区华通丰田汽车销售服务有限公司', '大连中升丰田汽车销售服务有限公司', '沈阳中晨诚隆汽车销售服务有限公司', '沈阳中晨先锋丰田汽车销售服务有限公司', '沈阳中晨诚隆汽车销售服务有限公司', '辽宁上辽汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '大连中升日产汽车销售服务有限公司', '沈阳南星行汽车销售服务有限公司', '丹东江通汽车销售服务有限公司', '沈阳金廊华林汽车销售服务有限公司', '抚顺上辽汽车销售服务有限公司', '沈阳中资华大汽车销售服务有限公司', '大连东升汽车销售服务有限公司', '沈阳中资华大汽车销售服务有限公司', '铁岭和成汽车销售服务有限公司', '辽宁鑫盛特汽车销售服务有限公司', '大连福友达汽车销售服务有限公司', '沈阳华宝汽车销售服务有限公司', '大连中升汇驰汽车销售服务有限公司', '营口大通汽车销售服务有限公司', '辽宁大地汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '辽宁上辽汽车销售服务有限公司', '爱威马(辽宁)汽车销售服务有限责任公司', '本溪创为汽车销售服务有限公司', '沈阳金廊雷克萨斯汽车销售服务有限公司', '辽宁盛世福丰汽车销售服务有限公司', '沈阳华宝汽车销售服务有限公司（BMW宝马店）', '锦州市欧亚汽车销售服务有限责任公司', '辽宁鑫汇通汽车销售服务有限公司', '营口君恒汽车销售服务有限公司', '沈阳运通宸宝汽车销售服务有限公司', '葫芦岛华宝汽车销售服务有限公司（BMW宝马店）', '辽阳英迪汽车销售服务有限公司', '营口君恒汽车销售服务有限公司', '鞍山和卓汽车销售服务有限公司', '大连亿霖汽车销售服务有限公司', '大连嘉和润诚汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳市宝晋汽车销售服务有限公司（BMW宝马店）', '辽宁路赛得安特汽车销售服务有限公司', '沈阳中晨诚隆汽车销售服务有限公司', '沈阳中晨东怡汽车销售服务有限公司', '大连中升宏达汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '大连骏德汽车销售服务有限公司', '大连中升之宝汽车销售服务有限公司（BMW宝马店）', '大连中升汽车销售服务有限公司', '爱威马(辽宁)汽车销售服务有限责任公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳喜悦汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '沈阳路丰源汽车销售服务有限公司', '沈阳宇宸汽车销售服务有限公司', '辽宁奥通汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '鞍山和卓汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳南光丰田汽车销售服务有限公司', '大连金廊华林汽车销售服务有限公司', '鞍山四隆之星汽车销售服务有限公司', '辽宁新世纪汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '沈阳中升仕豪汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '大连中升之宝汽车销售服务有限公司（BMW宝马店）', '辽宁奥通汽车销售服务有限公司', '大连裕迪汽车销售服务有限公司', '沈阳喜悦汽车销售服务有限公司', '辽宁顺亿达汽车销售服务有限公司', '辽宁奥通汽车销售服务有限公司', '大连中升汽车销售服务有限公司', '大连壁虎汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '大连华菱盛大汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '辽宁众兴汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '大连万海汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '大连汇华丰田汽车销售服务有限公司', '大连领航汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '铁岭天广吉祥汽车销售服务有限公司', '辽宁利丰星行汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '辽宁神州红旗汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '沈阳金廊雷克萨斯汽车销售服务有限公司', '沈阳中晨先锋丰田汽车销售服务有限公司', '沈阳中晨诚隆汽车销售服务有限公司', '辽宁上辽汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '大连中升日产汽车销售服务有限公司', '沈阳南星行汽车销售服务有限公司', '丹东江通汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '沈阳金廊华林汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '抚顺上辽汽车销售服务有限公司', '沈阳中资华大汽车销售服务有限公司', '大连东升汽车销售服务有限公司', '沈阳中资华大汽车销售服务有限公司', '铁岭和成汽车销售服务有限公司', '辽宁鑫盛特汽车销售服务有限公司', '大连福友达汽车销售服务有限公司', '沈阳华宝汽车销售服务有限公司', '大连中升汇驰汽车销售服务有限公司', '营口大通汽车销售服务有限公司', '辽宁大地汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '辽宁上辽汽车销售服务有限公司', '爱威马(辽宁)汽车销售服务有限责任公司', '本溪创为汽车销售服务有限公司', '沈阳金廊雷克萨斯汽车销售服务有限公司', '辽宁盛世福丰汽车销售服务有限公司', '沈阳华宝汽车销售服务有限公司（BMW宝马店）', '锦州市欧亚汽车销售服务有限责任公司', '辽宁鑫汇通汽车销售服务有限公司', '营口君恒汽车销售服务有限公司', '沈阳运通宸宝汽车销售服务有限公司', '葫芦岛华宝汽车销售服务有限公司（BMW宝马店）', '辽阳英迪汽车销售服务有限公司', '营口君恒汽车销售服务有限公司', '鞍山和卓汽车销售服务有限公司', '大连亿霖汽车销售服务有限公司', '大连嘉和润诚汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳市宝晋汽车销售服务有限公司（BMW宝马店）', '辽宁路赛得安特汽车销售服务有限公司', '沈阳中晨诚隆汽车销售服务有限公司', '沈阳中晨东怡汽车销售服务有限公司', '大连中升宏达汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '大连骏德汽车销售服务有限公司', '大连中升之宝汽车销售服务有限公司（BMW宝马店）', '大连中升汽车销售服务有限公司', '爱威马(辽宁)汽车销售服务有限责任公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳喜悦汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '沈阳路丰源汽车销售服务有限公司', '沈阳宇宸汽车销售服务有限公司', '辽宁奥通汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '鞍山和卓汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳南光丰田汽车销售服务有限公司', '大连金廊华林汽车销售服务有限公司', '鞍山四隆之星汽车销售服务有限公司', '辽宁新世纪汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '沈阳中升仕豪汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '大连中升之宝汽车销售服务有限公司（BMW宝马店）', '辽宁奥通汽车销售服务有限公司', '大连裕迪汽车销售服务有限公司', '沈阳喜悦汽车销售服务有限公司', '辽宁顺亿达汽车销售服务有限公司', '辽宁奥通汽车销售服务有限公司', '大连中升汽车销售服务有限公司', '大连壁虎汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '大连华菱盛大汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '辽宁众兴汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '大连万海汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '大连汇华丰田汽车销售服务有限公司', '大连领航汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '铁岭天广吉祥汽车销售服务有限公司', '辽宁利丰星行汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '辽宁神州红旗汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '沈阳中资华大汽车销售服务有限公司', '大连东升汽车销售服务有限公司', '沈阳中资华大汽车销售服务有限公司', '铁岭和成汽车销售服务有限公司', '辽宁鑫盛特汽车销售服务有限公司', '大连福友达汽车销售服务有限公司', '沈阳华宝汽车销售服务有限公司', '大连中升汇驰汽车销售服务有限公司', '营口大通汽车销售服务有限公司', '辽宁大地汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '辽宁上辽汽车销售服务有限公司', '爱威马(辽宁)汽车销售服务有限责任公司', '本溪创为汽车销售服务有限公司', '沈阳金廊雷克萨斯汽车销售服务有限公司', '辽宁盛世福丰汽车销售服务有限公司', '沈阳华宝汽车销售服务有限公司（BMW宝马店）', '锦州市欧亚汽车销售服务有限责任公司', '辽宁鑫汇通汽车销售服务有限公司', '营口君恒汽车销售服务有限公司', '沈阳运通宸宝汽车销售服务有限公司', '葫芦岛华宝汽车销售服务有限公司（BMW宝马店）', '辽阳英迪汽车销售服务有限公司', '营口君恒汽车销售服务有限公司', '鞍山和卓汽车销售服务有限公司', '大连亿霖汽车销售服务有限公司', '大连嘉和润诚汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳市宝晋汽车销售服务有限公司（BMW宝马店）', '辽宁路赛得安特汽车销售服务有限公司', '沈阳中晨诚隆汽车销售服务有限公司', '沈阳中晨东怡汽车销售服务有限公司', '大连中升宏达汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '大连骏德汽车销售服务有限公司', '大连中升之宝汽车销售服务有限公司（BMW宝马店）', '大连中升汽车销售服务有限公司', '爱威马(辽宁)汽车销售服务有限责任公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳喜悦汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '沈阳路丰源汽车销售服务有限公司', '沈阳宇宸汽车销售服务有限公司', '辽宁奥通汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '鞍山和卓汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳南光丰田汽车销售服务有限公司', '大连金廊华林汽车销售服务有限公司', '鞍山四隆之星汽车销售服务有限公司', '辽宁新世纪汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '沈阳中升仕豪汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '大连中升之宝汽车销售服务有限公司（BMW宝马店）', '辽宁奥通汽车销售服务有限公司', '大连裕迪汽车销售服务有限公司', '沈阳喜悦汽车销售服务有限公司', '辽宁顺亿达汽车销售服务有限公司', '辽宁奥通汽车销售服务有限公司', '大连中升汽车销售服务有限公司', '大连壁虎汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '大连华菱盛大汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '辽宁众兴汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '大连万海汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '大连汇华丰田汽车销售服务有限公司', '大连领航汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '铁岭天广吉祥汽车销售服务有限公司', '辽宁利丰星行汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '辽宁神州红旗汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '沈阳金廊雷克萨斯汽车销售服务有限公司', '沈阳华宝汽车销售服务有限公司', '大连中升汇驰汽车销售服务有限公司', '营口大通汽车销售服务有限公司', '辽宁大地汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '辽宁上辽汽车销售服务有限公司', '爱威马(辽宁)汽车销售服务有限责任公司', '本溪创为汽车销售服务有限公司', '沈阳金廊雷克萨斯汽车销售服务有限公司', '辽宁盛世福丰汽车销售服务有限公司', '沈阳华宝汽车销售服务有限公司（BMW宝马店）', '锦州市欧亚汽车销售服务有限责任公司', '辽宁鑫汇通汽车销售服务有限公司', '营口君恒汽车销售服务有限公司', '沈阳运通宸宝汽车销售服务有限公司', '葫芦岛华宝汽车销售服务有限公司（BMW宝马店）', '辽阳英迪汽车销售服务有限公司', '营口君恒汽车销售服务有限公司', '鞍山和卓汽车销售服务有限公司', '大连亿霖汽车销售服务有限公司', '大连嘉和润诚汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳市宝晋汽车销售服务有限公司（BMW宝马店）', '辽宁路赛得安特汽车销售服务有限公司', '沈阳中晨诚隆汽车销售服务有限公司', '沈阳中晨东怡汽车销售服务有限公司', '大连中升宏达汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '大连骏德汽车销售服务有限公司', '大连中升之宝汽车销售服务有限公司（BMW宝马店）', '大连中升汽车销售服务有限公司', '爱威马(辽宁)汽车销售服务有限责任公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳喜悦汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '沈阳路丰源汽车销售服务有限公司', '沈阳宇宸汽车销售服务有限公司', '辽宁奥通汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '鞍山和卓汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳南光丰田汽车销售服务有限公司', '大连金廊华林汽车销售服务有限公司', '鞍山四隆之星汽车销售服务有限公司', '辽宁新世纪汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '沈阳中升仕豪汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '大连中升之宝汽车销售服务有限公司（BMW宝马店）', '辽宁奥通汽车销售服务有限公司', '大连裕迪汽车销售服务有限公司', '沈阳喜悦汽车销售服务有限公司', '辽宁顺亿达汽车销售服务有限公司', '辽宁奥通汽车销售服务有限公司', '大连中升汽车销售服务有限公司', '大连壁虎汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '大连华菱盛大汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '辽宁众兴汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '大连万海汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '大连汇华丰田汽车销售服务有限公司', '大连领航汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '铁岭天广吉祥汽车销售服务有限公司', '辽宁利丰星行汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '辽宁神州红旗汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '沈阳金廊雷克萨斯汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳市宝晋汽车销售服务有限公司（BMW宝马店）', '辽宁路赛得安特汽车销售服务有限公司', '沈阳中晨诚隆汽车销售服务有限公司', '辽宁鑫汇通汽车销售服务有限公司', '大连中升宏达汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '大连骏德汽车销售服务有限公司', '大连中升之宝汽车销售服务有限公司（BMW宝马店）', '大连中升汽车销售服务有限公司', '爱威马(辽宁)汽车销售服务有限责任公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳喜悦汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '沈阳路丰源汽车销售服务有限公司', '沈阳宇宸汽车销售服务有限公司', '辽宁奥通汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '鞍山和卓汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳中晨东怡汽车销售服务有限公司', '沈阳南光丰田汽车销售服务有限公司', '大连金廊华林汽车销售服务有限公司', '鞍山四隆之星汽车销售服务有限公司', '辽宁新世纪汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '沈阳中升仕豪汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '大连中升之宝汽车销售服务有限公司（BMW宝马店）', '辽宁奥通汽车销售服务有限公司', '大连裕迪汽车销售服务有限公司', '沈阳喜悦汽车销售服务有限公司', '辽宁顺亿达汽车销售服务有限公司', '辽宁奥通汽车销售服务有限公司', '大连中升汽车销售服务有限公司', '大连壁虎汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '大连华菱盛大汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '辽宁众兴汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '大连万海汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '大连汇华丰田汽车销售服务有限公司', '大连领航汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '铁岭天广吉祥汽车销售服务有限公司', '辽宁利丰星行汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '辽宁神州红旗汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '沈阳金廊雷克萨斯汽车销售服务有限公司', '沈阳宇宸汽车销售服务有限公司', '辽宁奥通汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '鞍山和卓汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '沈阳南光丰田汽车销售服务有限公司', '大连金廊华林汽车销售服务有限公司', '鞍山四隆之星汽车销售服务有限公司', '辽宁新世纪汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '沈阳中升仕豪汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '大连中升之宝汽车销售服务有限公司（BMW宝马店）', '辽宁奥通汽车销售服务有限公司', '大连裕迪汽车销售服务有限公司', '沈阳喜悦汽车销售服务有限公司', '辽宁顺亿达汽车销售服务有限公司', '辽宁奥通汽车销售服务有限公司', '大连中升汽车销售服务有限公司', '大连壁虎汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '大连华菱盛大汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '辽宁众兴汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '大连万海汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '大连汇华丰田汽车销售服务有限公司', '大连领航汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '铁岭天广吉祥汽车销售服务有限公司', '辽宁利丰星行汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '辽宁神州红旗汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '沈阳金廊雷克萨斯汽车销售服务有限公司', '沈阳喜悦汽车销售服务有限公司', '辽宁顺亿达汽车销售服务有限公司', '辽宁奥通汽车销售服务有限公司', '大连中升汽车销售服务有限公司', '大连壁虎汽车销售服务有限公司', '辽宁兴锐汽车销售服务有限公司', '大连华菱盛大汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '辽宁中升捷通汽车销售服务有限公司', '辽宁众兴汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '大连万海汽车销售服务有限公司', '大连中升之星汽车销售服务有限公司', '大连汇华丰田汽车销售服务有限公司', '大连领航汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '铁岭天广吉祥汽车销售服务有限公司', '辽宁利丰星行汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '辽宁神州红旗汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '沈阳金廊雷克萨斯汽车销售服务有限公司', '大连双麟汽车销售服务有限公司', '铁岭天广吉祥汽车销售服务有限公司', '辽宁利丰星行汽车销售服务有限公司', '辽宁路赛得安特汽车销售服务有限公司', '辽宁神州红旗汽车销售服务有限公司', '同享宝时汽车销售服务有限公司', '沈阳金廊雷克萨斯汽车销售服务有限公司']


phone_num = []
for gongsi in gongsi_list:
    url = "https://www.qcc.com/web/search?key={}".format(gongsi)

    hed = {
        "host": "www.qcc.com",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.106 Safari/537.36",
        "accept": "application/json, text/plain, */*",
        "cookie": "acw_tc=7d27c71e16110277339175502eab199279c592245c7a0ff0e58e1a0ebb; QCCSESSID=15mtq0qqjjj3sf375pl5u8d8u3; UM_distinctid=17718bc6037acc-09488634218dab-31346d-144000-17718bc6038994; CNZZDATA1254842228=1175062772-1611026731-https%253A%252F%252Fwww.baidu.com%252F%7C1611026731; zg_did=%7B%22did%22%3A%20%2217718bc60f9193-0c13d5fe886c71-31346d-144000-17718bc60faa2%22%7D; hasShow=1; _uab_collina=161102773537012544932569; zg_de1d1a35bfa24ce29bbf2c7eb17e6c4f=%7B%22sid%22%3A%201611027734781%2C%22updated%22%3A%201611027803451%2C%22info%22%3A%201611027734784%2C%22superProperty%22%3A%20%22%7B%5C%22%E5%BA%94%E7%94%A8%E5%90%8D%E7%A7%B0%5C%22%3A%20%5C%22%E4%BC%81%E6%9F%A5%E6%9F%A5%E7%BD%91%E7%AB%99%5C%22%7D%22%2C%22platform%22%3A%20%22%7B%7D%22%2C%22utm%22%3A%20%22%7B%7D%22%2C%22referrerDomain%22%3A%20%22www.baidu.com%22%2C%22cuid%22%3A%20%224a0b39c609eb78b6b56632dd77e42ee4%22%7D"    }
    time.sleep(1)
    resp = requests.get(url = url,headers = hed).text
    # print(url)
    response = parsel.Selector(resp)
    # print(response)
    phone = response.xpath('//tr[@class="frtrt"]//span[@class="val"]//span/text()').get()
    # print(gongsi,phone)
    phone_num.append(phone)
    print(phone)


# # 添加 电话号
# import openpyxl
# #加载文bai件
# wb = openpyxl.load_workbook(r'C:\Users\liuyechun\Desktop\20210104晚确定的提高群内活跃度内容（岗位需求）用统计素材+联系方式.xlsx')
# #获得dusheet名称
# sheetNames = wb.sheetnames
# print(sheetNames)
# #sheetName1 = sheetNames[0]
# #根据名称获取第zhi一个sheet
# #sheet1 = wb[sheetName1]
# #根据索引获得第一个sheet
# sheet1 = wb.worksheets[0]


#excel中单元格为B2开始，即shu第2列，第2行

# for i in range(len(phone_num)):
#   sheet1.cell(i+2, 5).value=phone_num[i]
# #保存数据，如果提示权限错误，需要关闭打开的excel
# wb.save(r'C:\Users\liuyechun\Desktop\20210104晚确定的提高群内活跃度内容（岗位需求）用统计素材+联系方式demo.xlsx')





























































