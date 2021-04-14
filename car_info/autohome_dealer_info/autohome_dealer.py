




# 获取字段 ：品牌、地址、电话、店名


import requests.adapters
from requests.adapters import HTTPAdapter
import pypinyin
import parsel
from openpyxl import Workbook

"""全国"""

# wb = Workbook()
# ws = wb.active
# ws.append(['品牌','店名','电话','地址','城市'])
#
# # 解决ssl证书警告
# requests.packages.urllib3.disable_warnings()
#
# # 解决超过最大链接
# srequest = requests.session()
# srequest.mount('https://', HTTPAdapter(max_retries=10))
#
# # 解决重连
# srequest.keep_alive = False
# srequest.adapters.DEFAULT_RETRIES = 10
#
# headers = {
#     'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.106 Safari/537.36',
#     'Connection':'close',
#     'cookie':'vlid=1598334632611x7bYbNr3JN; sessionid=BB0564B9-7B92-4628-A03D-6DE5063240D1%7C%7C2020-08-25+13%3A50%3A32.313%7C%7Cwww.baidu.com; autoid=edb260ee2f8896bcbe87a8bec00502e1; area=210106; ahpau=1; sessionuid=BB0564B9-7B92-4628-A03D-6DE5063240D1%7C%7C2020-08-25+13%3A50%3A32.313%7C%7Cwww.baidu.com; __ah_uuid_ng=c_BB0564B9-7B92-4628-A03D-6DE5063240D1; Hm_lvt_9924a05a5a75caf05dbbfb51af638b07=1598335812; jrsfvi=1598424705503IWzQEODWWC%7Cwww.autohome.com.cn%7C6841722; orderCount=eyJzaXRlSWQiOiI1MSIsImNhdGVnb3J5SWQiOiI4MjEiLCJzdWJDYXRlZ29yeUlkIjoiMTU0MjAiLCJ1c2VySWQiOiIiLCJwdmlkQ2hhaW4iOiIxMDE1OTQsMTAxNTk0LDEwMTU5NCwxMDE1OTQsNjg0MTcyMiIsImFjY2Vzc1R5cGUiOiIxIiwiYXBwS2V5IjoiIiwibG9jQ2l0eUlkIjoiMjEwMTAwIiwibG9jUHJvdmluY2VJZCI6IjIxMDAwMCIsImRldmljZUlkIjoiIiwibG9hZElkIjoiMTU5ODQyNDcwNTUwM0lXelFFT0RXV0MiLCJzZXNzaW9uSWQiOiJCQjA1NjRCOS03QjkyLTQ2MjgtQTAzRC02REU1MDYzMjQwRDF8fDIwMjAtMDgtMjUrMTM6NTA6MzIuMzEzfHx3d3cuYmFpZHUuY29tIiwidmlzaXRfaW5mbyI6IkJCMDU2NEI5LTdCOTItNDYyOC1BMDNELTZERTUwNjMyNDBEMXx8OTlFRDk2OTktQ0YzOC00MThBLUE0QTYtMTU1MUQxNERGRDg4fHwyMDIwMDUwOXx8MDF8fDYiLCJjdXJQdmFyZWFJZCI6IjY4NDE3MjIiLCJwdmFyZWFJZCI6IjY4NDE3MjIifQ==; Hm_lpvt_9924a05a5a75caf05dbbfb51af638b07=1598515794; ahsids=874_4745_5009_4105; FootPrints=45048%7C2020-8-27%2C44696%7C2020-8-27%2C37352%7C2020-8-27%2C38170%7C2020-8-27%2C45453%7C2020-8-27%2C; ahplid=1599162565800; ahpvno=94; sessionip=119.119.130.86; v_no=1; visit_info_ad=BB0564B9-7B92-4628-A03D-6DE5063240D1||58D95981-AAD3-41AA-A6C4-4824CC71D019||-1||-1||1; ref=www.baidu.com%7C0%7C0%7C0%7C2020-08-28+09%3A06%3A35.007%7C2020-08-25+13%3A50%3A32.313; sessionvid=58D95981-AAD3-41AA-A6C4-4824CC71D019; ahrlid=1598576794896Aidc0tZHac-1598576797201'
# }
# proxies = {  # 代理ip
#     "https": "https://221.122.91.61:9400"
# }
#
# address_list = ['沈阳','大连','鞍山','抚顺','本溪','丹东','锦州','营口','阜新','辽阳','盘锦','铁岭','chaoyang','葫芦岛']
# for add in address_list:
#     s = ''
#     for py in pypinyin.pinyin(add, style=pypinyin.NORMAL):
#         s += ''.join(py)
#     for page in range(1,13):
#         url = 'https://dealer.autohome.com.cn/{}/0/0/0/0/{}/1/0/0.html'.format(s,page)
#         print("请求的url",url)
#         resp_html = srequest.get(url,headers=headers)
#         resp_data = parsel.Selector(resp_html.text)
#
#
#
#         data_list = resp_data.xpath('//ul[@class="list-box"]/li[@class="list-item"]')
#
#         for li in data_list:
#             # 获取品牌
#             brand_name = li.xpath('./ul[@class="info-wrap"]/li[2]/span/em/text()').get()
#             print('品牌', brand_name)
#
#             # 获取店名
#             store_name = li.xpath('./ul[@class="info-wrap"]/li[1]/a/span/text()').get()
#             print('店名',store_name)
#
#
#             # 电话号
#             telephone_number = li.xpath('./ul[@class="info-wrap"]/li[3]/span[@class="tel"]/text()').get()
#             print('电话',telephone_number)
#
#             # 获取地址
#             address_of_4s = li.xpath('./ul[@class="info-wrap"]/li[4]/span[@class="info-addr"]/text()').get()
#             print('地址',address_of_4s)
#             if '重庆' in address_of_4s:
#                 continue
#             if '广州' in address_of_4s:
#                 continue
#             print('++++++++++++++++++++++++++++++++++++分割线++++++++++++++++++++++++++++++++++++')
#
#             # 获取城市
#             city_name = resp_data.xpath('//span[@id="change-city"]/span[@class="city-now"]/text()').get()
#             print('城市', city_name)
#
#             # Excel 存储
#             line = [brand_name, store_name, telephone_number,address_of_4s,city_name]
#             ws.append(line)
#             wb.save(r'C:\Users\liuyechun\Desktop\汽车之家.xlsx')


"""吉林"""

# wb = Workbook()
# ws = wb.active
# ws.append(['品牌','店名','电话','地址','城市'])
#
# # 解决ssl证书警告
# requests.packages.urllib3.disable_warnings()
#
# # 解决超过最大链接
# srequest = requests.session()
# srequest.mount('https://', HTTPAdapter(max_retries=10))
#
# # 解决重连
# srequest.keep_alive = False
# srequest.adapters.DEFAULT_RETRIES = 10
#
# headers = {
#     'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.106 Safari/537.36',
#     'Connection':'close',
#     'cookie':'vlid=1598334632611x7bYbNr3JN; sessionid=BB0564B9-7B92-4628-A03D-6DE5063240D1%7C%7C2020-08-25+13%3A50%3A32.313%7C%7Cwww.baidu.com; autoid=edb260ee2f8896bcbe87a8bec00502e1; area=210106; ahpau=1; sessionuid=BB0564B9-7B92-4628-A03D-6DE5063240D1%7C%7C2020-08-25+13%3A50%3A32.313%7C%7Cwww.baidu.com; __ah_uuid_ng=c_BB0564B9-7B92-4628-A03D-6DE5063240D1; Hm_lvt_9924a05a5a75caf05dbbfb51af638b07=1598335812; jrsfvi=1598424705503IWzQEODWWC%7Cwww.autohome.com.cn%7C6841722; orderCount=eyJzaXRlSWQiOiI1MSIsImNhdGVnb3J5SWQiOiI4MjEiLCJzdWJDYXRlZ29yeUlkIjoiMTU0MjAiLCJ1c2VySWQiOiIiLCJwdmlkQ2hhaW4iOiIxMDE1OTQsMTAxNTk0LDEwMTU5NCwxMDE1OTQsNjg0MTcyMiIsImFjY2Vzc1R5cGUiOiIxIiwiYXBwS2V5IjoiIiwibG9jQ2l0eUlkIjoiMjEwMTAwIiwibG9jUHJvdmluY2VJZCI6IjIxMDAwMCIsImRldmljZUlkIjoiIiwibG9hZElkIjoiMTU5ODQyNDcwNTUwM0lXelFFT0RXV0MiLCJzZXNzaW9uSWQiOiJCQjA1NjRCOS03QjkyLTQ2MjgtQTAzRC02REU1MDYzMjQwRDF8fDIwMjAtMDgtMjUrMTM6NTA6MzIuMzEzfHx3d3cuYmFpZHUuY29tIiwidmlzaXRfaW5mbyI6IkJCMDU2NEI5LTdCOTItNDYyOC1BMDNELTZERTUwNjMyNDBEMXx8OTlFRDk2OTktQ0YzOC00MThBLUE0QTYtMTU1MUQxNERGRDg4fHwyMDIwMDUwOXx8MDF8fDYiLCJjdXJQdmFyZWFJZCI6IjY4NDE3MjIiLCJwdmFyZWFJZCI6IjY4NDE3MjIifQ==; Hm_lpvt_9924a05a5a75caf05dbbfb51af638b07=1598515794; ahsids=874_4745_5009_4105; FootPrints=45048%7C2020-8-27%2C44696%7C2020-8-27%2C37352%7C2020-8-27%2C38170%7C2020-8-27%2C45453%7C2020-8-27%2C; ahplid=1599162565800; ahpvno=94; sessionip=119.119.130.86; v_no=1; visit_info_ad=BB0564B9-7B92-4628-A03D-6DE5063240D1||58D95981-AAD3-41AA-A6C4-4824CC71D019||-1||-1||1; ref=www.baidu.com%7C0%7C0%7C0%7C2020-08-28+09%3A06%3A35.007%7C2020-08-25+13%3A50%3A32.313; sessionvid=58D95981-AAD3-41AA-A6C4-4824CC71D019; ahrlid=1598576794896Aidc0tZHac-1598576797201'
# }
# proxies = {  # 代理ip
#     "https": "https://221.122.91.61:9400"
# }
#
# address_list = ['长春','吉林市','四平','辽源','通化','白山','松原','白城','延边']
# for add in address_list:
#     s = ''
#     for py in pypinyin.pinyin(add, style=pypinyin.NORMAL):
#         s += ''.join(py)
#     for page in range(1,13):
#         url = 'https://dealer.autohome.com.cn/{}/0/0/0/0/{}/1/0/0.html'.format(s,page)
#         print("请求的url",url)
#         resp_html = srequest.get(url,headers=headers)
#         resp_html.encoding = 'gbk'
#         resp_data = parsel.Selector(resp_html.text)
#
#         data_list = resp_data.xpath('//ul[@class="list-box"]/li[@class="list-item"]')
#
#         for li in data_list:
#             # 获取品牌
#             brand_name = li.xpath('./ul[@class="info-wrap"]/li[2]/span/em/text()').get()
#             print('品牌', brand_name)
#
#             # 获取店名
#             store_name = li.xpath('./ul[@class="info-wrap"]/li[1]/a/span/text()').get()
#             print('店名',store_name)
#
#
#             # 电话号
#             telephone_number = li.xpath('./ul[@class="info-wrap"]/li[3]/span[@class="tel"]/text()').get()
#             print('电话',telephone_number)
#
#             # 获取地址
#             address_of_4s = li.xpath('./ul[@class="info-wrap"]/li[4]/span[@class="info-addr"]/text()').get()
#             print('地址',address_of_4s)
#             if '重庆' in address_of_4s:
#                 continue
#             if '广州' in address_of_4s:
#                 continue
#             print('++++++++++++++++++++++++++++++++++++分割线++++++++++++++++++++++++++++++++++++')
#
#             # 获取城市
#             city_name = resp_data.xpath('//span[@id="change-city"]/span[@class="city-now"]/text()').get()
#             print('城市', city_name)
#
#             # Excel 存储
#             line = [brand_name, store_name, telephone_number,address_of_4s,city_name]
#             ws.append(line)
#             wb.save(r'C:\Users\liuyechun\Desktop\汽车之家-吉林.xlsx')


"""黑龙江"""
#
# wb = Workbook()
# ws = wb.active
# ws.append(['序号','品牌','店名','电话','地址','城市'])
#
#
# # 解决ssl证书警告
# requests.packages.urllib3.disable_warnings()
#
# # 解决超过最大链接
# srequest = requests.session()
# srequest.mount('https://', HTTPAdapter(max_retries=10))
#
# # 解决重连
# srequest.keep_alive = False
# srequest.adapters.DEFAULT_RETRIES = 10
#
# headers = {
#     'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.106 Safari/537.36',
#     'Connection':'close',
#     'cookie':'vlid=1598334632611x7bYbNr3JN; sessionid=BB0564B9-7B92-4628-A03D-6DE5063240D1%7C%7C2020-08-25+13%3A50%3A32.313%7C%7Cwww.baidu.com; autoid=edb260ee2f8896bcbe87a8bec00502e1; area=210106; ahpau=1; sessionuid=BB0564B9-7B92-4628-A03D-6DE5063240D1%7C%7C2020-08-25+13%3A50%3A32.313%7C%7Cwww.baidu.com; __ah_uuid_ng=c_BB0564B9-7B92-4628-A03D-6DE5063240D1; Hm_lvt_9924a05a5a75caf05dbbfb51af638b07=1598335812; jrsfvi=1598424705503IWzQEODWWC%7Cwww.autohome.com.cn%7C6841722; orderCount=eyJzaXRlSWQiOiI1MSIsImNhdGVnb3J5SWQiOiI4MjEiLCJzdWJDYXRlZ29yeUlkIjoiMTU0MjAiLCJ1c2VySWQiOiIiLCJwdmlkQ2hhaW4iOiIxMDE1OTQsMTAxNTk0LDEwMTU5NCwxMDE1OTQsNjg0MTcyMiIsImFjY2Vzc1R5cGUiOiIxIiwiYXBwS2V5IjoiIiwibG9jQ2l0eUlkIjoiMjEwMTAwIiwibG9jUHJvdmluY2VJZCI6IjIxMDAwMCIsImRldmljZUlkIjoiIiwibG9hZElkIjoiMTU5ODQyNDcwNTUwM0lXelFFT0RXV0MiLCJzZXNzaW9uSWQiOiJCQjA1NjRCOS03QjkyLTQ2MjgtQTAzRC02REU1MDYzMjQwRDF8fDIwMjAtMDgtMjUrMTM6NTA6MzIuMzEzfHx3d3cuYmFpZHUuY29tIiwidmlzaXRfaW5mbyI6IkJCMDU2NEI5LTdCOTItNDYyOC1BMDNELTZERTUwNjMyNDBEMXx8OTlFRDk2OTktQ0YzOC00MThBLUE0QTYtMTU1MUQxNERGRDg4fHwyMDIwMDUwOXx8MDF8fDYiLCJjdXJQdmFyZWFJZCI6IjY4NDE3MjIiLCJwdmFyZWFJZCI6IjY4NDE3MjIifQ==; Hm_lpvt_9924a05a5a75caf05dbbfb51af638b07=1598515794; ahsids=874_4745_5009_4105; FootPrints=45048%7C2020-8-27%2C44696%7C2020-8-27%2C37352%7C2020-8-27%2C38170%7C2020-8-27%2C45453%7C2020-8-27%2C; ahplid=1599162565800; ahpvno=94; sessionip=119.119.130.86; v_no=1; visit_info_ad=BB0564B9-7B92-4628-A03D-6DE5063240D1||58D95981-AAD3-41AA-A6C4-4824CC71D019||-1||-1||1; ref=www.baidu.com%7C0%7C0%7C0%7C2020-08-28+09%3A06%3A35.007%7C2020-08-25+13%3A50%3A32.313; sessionvid=58D95981-AAD3-41AA-A6C4-4824CC71D019; ahrlid=1598576794896Aidc0tZHac-1598576797201'
# }
# proxies = {  # 代理ip
#     "https": "https://221.122.91.61:9400"
# }
#
# address_list = ['哈尔滨','齐齐哈尔','鸡西','鹤岗','双鸭山','大庆','伊春','佳木斯','七台河','牡丹江','黑河','绥化','大兴安岭']
# for add in address_list:
#     s = ''
#     for py in pypinyin.pinyin(add, style=pypinyin.NORMAL):
#         s += ''.join(py)
#     for page in range(1,13):
#         url = 'https://dealer.autohome.com.cn/{}/0/0/0/0/{}/1/0/0.html'.format(s,page)
#         print("请求的url",url)
#         resp_html = srequest.get(url,headers=headers)
#         resp_html.encoding = 'gbk'
#         resp_data = parsel.Selector(resp_html.text)
#
#         data_list = resp_data.xpath('//ul[@class="list-box"]/li[@class="list-item"]')
#
#         for li in data_list:
#             # 获取品牌
#             brand_name = li.xpath('./ul[@class="info-wrap"]/li[2]/span/em/text()').get()
#             print('品牌', brand_name)
#
#             # 获取店名
#             store_name = li.xpath('./ul[@class="info-wrap"]/li[1]/a/span/text()').get()
#             print('店名',store_name)
#
#
#             # 电话号
#             telephone_number = li.xpath('./ul[@class="info-wrap"]/li[3]/span[@class="tel"]/text()').get()
#             print('电话',telephone_number)
#
#             # 获取地址
#             address_of_4s = li.xpath('./ul[@class="info-wrap"]/li[4]/span[@class="info-addr"]/text()').get()
#             if '重庆' in address_of_4s:
#                 continue
#             if '广州' in address_of_4s:
#                 continue
#             print('地址', address_of_4s)
#             print('++++++++++++++++++++++++++++++++++++分割线++++++++++++++++++++++++++++++++++++')
#
#             # 获取城市
#             city_name = resp_data.xpath('//span[@id="change-city"]/span[@class="city-now"]/text()').get()
#             print('城市', city_name)
#
#             # Excel 存储
#             line = [brand_name, store_name, telephone_number,address_of_4s,city_name]
#             ws.append(line)
#             wb.save(r'C:\Users\liuyechun\Desktop\汽车之家-黑龙江.xlsx')



"""山东"""

wb = Workbook()
ws = wb.active
ws.append(['品牌','店名','电话','地址','城市'])


# 解决ssl证书警告
requests.packages.urllib3.disable_warnings()

# 解决超过最大链接
srequest = requests.session()
srequest.mount('https://', HTTPAdapter(max_retries=10))

# 解决重连
srequest.keep_alive = False
srequest.adapters.DEFAULT_RETRIES = 10

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.106 Safari/537.36',
    'Connection':'close',
    'cookie':'vlid=1598334632611x7bYbNr3JN; sessionid=BB0564B9-7B92-4628-A03D-6DE5063240D1%7C%7C2020-08-25+13%3A50%3A32.313%7C%7Cwww.baidu.com; autoid=edb260ee2f8896bcbe87a8bec00502e1; area=210106; ahpau=1; sessionuid=BB0564B9-7B92-4628-A03D-6DE5063240D1%7C%7C2020-08-25+13%3A50%3A32.313%7C%7Cwww.baidu.com; __ah_uuid_ng=c_BB0564B9-7B92-4628-A03D-6DE5063240D1; Hm_lvt_9924a05a5a75caf05dbbfb51af638b07=1598335812; jrsfvi=1598424705503IWzQEODWWC%7Cwww.autohome.com.cn%7C6841722; orderCount=eyJzaXRlSWQiOiI1MSIsImNhdGVnb3J5SWQiOiI4MjEiLCJzdWJDYXRlZ29yeUlkIjoiMTU0MjAiLCJ1c2VySWQiOiIiLCJwdmlkQ2hhaW4iOiIxMDE1OTQsMTAxNTk0LDEwMTU5NCwxMDE1OTQsNjg0MTcyMiIsImFjY2Vzc1R5cGUiOiIxIiwiYXBwS2V5IjoiIiwibG9jQ2l0eUlkIjoiMjEwMTAwIiwibG9jUHJvdmluY2VJZCI6IjIxMDAwMCIsImRldmljZUlkIjoiIiwibG9hZElkIjoiMTU5ODQyNDcwNTUwM0lXelFFT0RXV0MiLCJzZXNzaW9uSWQiOiJCQjA1NjRCOS03QjkyLTQ2MjgtQTAzRC02REU1MDYzMjQwRDF8fDIwMjAtMDgtMjUrMTM6NTA6MzIuMzEzfHx3d3cuYmFpZHUuY29tIiwidmlzaXRfaW5mbyI6IkJCMDU2NEI5LTdCOTItNDYyOC1BMDNELTZERTUwNjMyNDBEMXx8OTlFRDk2OTktQ0YzOC00MThBLUE0QTYtMTU1MUQxNERGRDg4fHwyMDIwMDUwOXx8MDF8fDYiLCJjdXJQdmFyZWFJZCI6IjY4NDE3MjIiLCJwdmFyZWFJZCI6IjY4NDE3MjIifQ==; Hm_lpvt_9924a05a5a75caf05dbbfb51af638b07=1598515794; ahsids=874_4745_5009_4105; FootPrints=45048%7C2020-8-27%2C44696%7C2020-8-27%2C37352%7C2020-8-27%2C38170%7C2020-8-27%2C45453%7C2020-8-27%2C; ahplid=1599162565800; ahpvno=94; sessionip=119.119.130.86; v_no=1; visit_info_ad=BB0564B9-7B92-4628-A03D-6DE5063240D1||58D95981-AAD3-41AA-A6C4-4824CC71D019||-1||-1||1; ref=www.baidu.com%7C0%7C0%7C0%7C2020-08-28+09%3A06%3A35.007%7C2020-08-25+13%3A50%3A32.313; sessionvid=58D95981-AAD3-41AA-A6C4-4824CC71D019; ahrlid=1598576794896Aidc0tZHac-1598576797201'
}
proxies = {  # 代理ip
    "https": "https://221.122.91.61:9400"
}

address_list = ['青岛','淄博','枣庄','东营','烟台','潍坊','济宁','泰安','威海','日照','莱芜','临沂','德州','聊城','滨州','菏泽']

for add in address_list:
    s = ''
    for py in pypinyin.pinyin(add, style=pypinyin.NORMAL):
        s += ''.join(py)
    for page in range(1,13):
        url = 'https://dealer.autohome.com.cn/{}/0/0/0/0/{}/1/0/0.html'.format(s,page)
        print("请求的url",url)
        resp_html = srequest.get(url,headers=headers)
        resp_html.encoding = 'gbk'
        resp_data = parsel.Selector(resp_html.text)

        data_list = resp_data.xpath('//ul[@class="list-box"]/li[@class="list-item"]')

        for li in data_list:
            # 获取品牌
            brand_name = li.xpath('./ul[@class="info-wrap"]/li[2]/span/em/text()').get()
            print('品牌', brand_name)

            # 获取店名
            store_name = li.xpath('./ul[@class="info-wrap"]/li[1]/a/span/text()').get()
            print('店名',store_name)


            # 电话号
            telephone_number = li.xpath('./ul[@class="info-wrap"]/li[3]/span[@class="tel"]/text()').get()
            print('电话',telephone_number)

            # 获取地址
            address_of_4s = li.xpath('./ul[@class="info-wrap"]/li[4]/span[@class="info-addr"]/text()').get()
            if '重庆' in address_of_4s:
                continue
            if '广州' in address_of_4s:
                continue
            print('地址', address_of_4s)
            print('++++++++++++++++++++++++++++++++++++分割线++++++++++++++++++++++++++++++++++++')

            # 获取城市
            city_name = resp_data.xpath('//span[@id="change-city"]/span[@class="city-now"]/text()').get()
            print('城市', city_name)

            # Excel 存储
            line = [brand_name, store_name, telephone_number,address_of_4s,city_name]
            ws.append(line)
            wb.save(r'C:\Users\liuyechun\Desktop\汽车之家-山东.xlsx')

"""山西"""
import requests.adapters
from requests.adapters import HTTPAdapter
import pypinyin
import parsel
from openpyxl import Workbook

# wb = Workbook()
# ws = wb.active
# ws.append(['序号','品牌','店名','电话','地址','城市'])
#
#
# # 解决ssl证书警告
# requests.packages.urllib3.disable_warnings()
#
# # 解决超过最大链接
# srequest = requests.session()
# srequest.mount('https://', HTTPAdapter(max_retries=10))
#
# # 解决重连
# srequest.keep_alive = False
# srequest.adapters.DEFAULT_RETRIES = 10
#
# headers = {
#     'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.106 Safari/537.36',
#     'Connection':'close',
#     'cookie':'vlid=1598334632611x7bYbNr3JN; sessionid=BB0564B9-7B92-4628-A03D-6DE5063240D1%7C%7C2020-08-25+13%3A50%3A32.313%7C%7Cwww.baidu.com; autoid=edb260ee2f8896bcbe87a8bec00502e1; area=210106; ahpau=1; sessionuid=BB0564B9-7B92-4628-A03D-6DE5063240D1%7C%7C2020-08-25+13%3A50%3A32.313%7C%7Cwww.baidu.com; __ah_uuid_ng=c_BB0564B9-7B92-4628-A03D-6DE5063240D1; Hm_lvt_9924a05a5a75caf05dbbfb51af638b07=1598335812; jrsfvi=1598424705503IWzQEODWWC%7Cwww.autohome.com.cn%7C6841722; orderCount=eyJzaXRlSWQiOiI1MSIsImNhdGVnb3J5SWQiOiI4MjEiLCJzdWJDYXRlZ29yeUlkIjoiMTU0MjAiLCJ1c2VySWQiOiIiLCJwdmlkQ2hhaW4iOiIxMDE1OTQsMTAxNTk0LDEwMTU5NCwxMDE1OTQsNjg0MTcyMiIsImFjY2Vzc1R5cGUiOiIxIiwiYXBwS2V5IjoiIiwibG9jQ2l0eUlkIjoiMjEwMTAwIiwibG9jUHJvdmluY2VJZCI6IjIxMDAwMCIsImRldmljZUlkIjoiIiwibG9hZElkIjoiMTU5ODQyNDcwNTUwM0lXelFFT0RXV0MiLCJzZXNzaW9uSWQiOiJCQjA1NjRCOS03QjkyLTQ2MjgtQTAzRC02REU1MDYzMjQwRDF8fDIwMjAtMDgtMjUrMTM6NTA6MzIuMzEzfHx3d3cuYmFpZHUuY29tIiwidmlzaXRfaW5mbyI6IkJCMDU2NEI5LTdCOTItNDYyOC1BMDNELTZERTUwNjMyNDBEMXx8OTlFRDk2OTktQ0YzOC00MThBLUE0QTYtMTU1MUQxNERGRDg4fHwyMDIwMDUwOXx8MDF8fDYiLCJjdXJQdmFyZWFJZCI6IjY4NDE3MjIiLCJwdmFyZWFJZCI6IjY4NDE3MjIifQ==; Hm_lpvt_9924a05a5a75caf05dbbfb51af638b07=1598515794; ahsids=874_4745_5009_4105; FootPrints=45048%7C2020-8-27%2C44696%7C2020-8-27%2C37352%7C2020-8-27%2C38170%7C2020-8-27%2C45453%7C2020-8-27%2C; ahplid=1599162565800; ahpvno=94; sessionip=119.119.130.86; v_no=1; visit_info_ad=BB0564B9-7B92-4628-A03D-6DE5063240D1||58D95981-AAD3-41AA-A6C4-4824CC71D019||-1||-1||1; ref=www.baidu.com%7C0%7C0%7C0%7C2020-08-28+09%3A06%3A35.007%7C2020-08-25+13%3A50%3A32.313; sessionvid=58D95981-AAD3-41AA-A6C4-4824CC71D019; ahrlid=1598576794896Aidc0tZHac-1598576797201'
# }
# proxies = {  # 代理ip
#     "https": "https://221.122.91.61:9400"
# }
#
# address_list = ['太原','大同','阳泉','长治','晋城','朔州','晋中','运城','忻州','临汾','吕梁']
# for add in address_list:
#     s = ''
#     for py in pypinyin.pinyin(add, style=pypinyin.NORMAL):
#         s += ''.join(py)
#     for page in range(1,13):
#         url = 'https://dealer.autohome.com.cn/{}/0/0/0/0/{}/1/0/0.html'.format(s,page)
#         print("请求的url",url)
#         resp_html = srequest.get(url,headers=headers)
#         resp_html.encoding = 'gbk'
#         resp_data = parsel.Selector(resp_html.text)
#
#         data_list = resp_data.xpath('//ul[@class="list-box"]/li[@class="list-item"]')
#
#         for li in data_list:
#             # 获取品牌
#             brand_name = li.xpath('./ul[@class="info-wrap"]/li[2]/span/em/text()').get()
#             print('品牌', brand_name)
#
#             # 获取店名
#             store_name = li.xpath('./ul[@class="info-wrap"]/li[1]/a/span/text()').get()
#             print('店名',store_name)
#
#
#             # 电话号
#             telephone_number = li.xpath('./ul[@class="info-wrap"]/li[3]/span[@class="tel"]/text()').get()
#             print('电话',telephone_number)
#
#             # 获取地址
#             address_of_4s = li.xpath('./ul[@class="info-wrap"]/li[4]/span[@class="info-addr"]/text()').get()
#             if '重庆' in address_of_4s:
#                 continue
#             if '广州' in address_of_4s:
#                 continue
#             print('地址', address_of_4s)
#             print('++++++++++++++++++++++++++++++++++++分割线++++++++++++++++++++++++++++++++++++')
#
#             # 获取城市
#             city_name = resp_data.xpath('//span[@id="change-city"]/span[@class="city-now"]/text()').get()
#             print('城市', city_name)
#
#             # Excel 存储
#             line = [brand_name, store_name, telephone_number,address_of_4s,city_name]
#             ws.append(line)
#             wb.save(r'C:\Users\liuyechun\Desktop\汽车之家-山西.xlsx')




"""河北"""
import requests.adapters
from requests.adapters import HTTPAdapter
import pypinyin
import parsel
from openpyxl import Workbook
# wb = Workbook()
# ws = wb.active
# ws.append(['序号','品牌','店名','电话','地址','城市'])
#
#
# # 解决ssl证书警告
# requests.packages.urllib3.disable_warnings()
#
# # 解决超过最大链接
# srequest = requests.session()
# srequest.mount('https://', HTTPAdapter(max_retries=10))
#
# # 解决重连
# srequest.keep_alive = False
# srequest.adapters.DEFAULT_RETRIES = 10
#
# headers = {
#     'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.106 Safari/537.36',
#     'Connection':'close',
#     'cookie':'vlid=1598334632611x7bYbNr3JN; sessionid=BB0564B9-7B92-4628-A03D-6DE5063240D1%7C%7C2020-08-25+13%3A50%3A32.313%7C%7Cwww.baidu.com; autoid=edb260ee2f8896bcbe87a8bec00502e1; area=210106; ahpau=1; sessionuid=BB0564B9-7B92-4628-A03D-6DE5063240D1%7C%7C2020-08-25+13%3A50%3A32.313%7C%7Cwww.baidu.com; __ah_uuid_ng=c_BB0564B9-7B92-4628-A03D-6DE5063240D1; Hm_lvt_9924a05a5a75caf05dbbfb51af638b07=1598335812; jrsfvi=1598424705503IWzQEODWWC%7Cwww.autohome.com.cn%7C6841722; orderCount=eyJzaXRlSWQiOiI1MSIsImNhdGVnb3J5SWQiOiI4MjEiLCJzdWJDYXRlZ29yeUlkIjoiMTU0MjAiLCJ1c2VySWQiOiIiLCJwdmlkQ2hhaW4iOiIxMDE1OTQsMTAxNTk0LDEwMTU5NCwxMDE1OTQsNjg0MTcyMiIsImFjY2Vzc1R5cGUiOiIxIiwiYXBwS2V5IjoiIiwibG9jQ2l0eUlkIjoiMjEwMTAwIiwibG9jUHJvdmluY2VJZCI6IjIxMDAwMCIsImRldmljZUlkIjoiIiwibG9hZElkIjoiMTU5ODQyNDcwNTUwM0lXelFFT0RXV0MiLCJzZXNzaW9uSWQiOiJCQjA1NjRCOS03QjkyLTQ2MjgtQTAzRC02REU1MDYzMjQwRDF8fDIwMjAtMDgtMjUrMTM6NTA6MzIuMzEzfHx3d3cuYmFpZHUuY29tIiwidmlzaXRfaW5mbyI6IkJCMDU2NEI5LTdCOTItNDYyOC1BMDNELTZERTUwNjMyNDBEMXx8OTlFRDk2OTktQ0YzOC00MThBLUE0QTYtMTU1MUQxNERGRDg4fHwyMDIwMDUwOXx8MDF8fDYiLCJjdXJQdmFyZWFJZCI6IjY4NDE3MjIiLCJwdmFyZWFJZCI6IjY4NDE3MjIifQ==; Hm_lpvt_9924a05a5a75caf05dbbfb51af638b07=1598515794; ahsids=874_4745_5009_4105; FootPrints=45048%7C2020-8-27%2C44696%7C2020-8-27%2C37352%7C2020-8-27%2C38170%7C2020-8-27%2C45453%7C2020-8-27%2C; ahplid=1599162565800; ahpvno=94; sessionip=119.119.130.86; v_no=1; visit_info_ad=BB0564B9-7B92-4628-A03D-6DE5063240D1||58D95981-AAD3-41AA-A6C4-4824CC71D019||-1||-1||1; ref=www.baidu.com%7C0%7C0%7C0%7C2020-08-28+09%3A06%3A35.007%7C2020-08-25+13%3A50%3A32.313; sessionvid=58D95981-AAD3-41AA-A6C4-4824CC71D019; ahrlid=1598576794896Aidc0tZHac-1598576797201'
# }
# proxies = {  # 代理ip
#     "https": "https://221.122.91.61:9400"
# }
#
# address_list = ['石家庄','唐山','秦皇岛','邯郸','邢台','保定','张家口','承德','沧州','廊坊','衡水']
# for add in address_list:
#     s = ''
#     for py in pypinyin.pinyin(add, style=pypinyin.NORMAL):
#         s += ''.join(py)
#     for page in range(1,13):
#         url = 'https://dealer.autohome.com.cn/{}/0/0/0/0/{}/1/0/0.html'.format(s,page)
#         print("请求的url",url)
#         resp_html = srequest.get(url,headers=headers)
#         resp_html.encoding = 'gbk'
#         resp_data = parsel.Selector(resp_html.text)
#
#         data_list = resp_data.xpath('//ul[@class="list-box"]/li[@class="list-item"]')
#
#         for li in data_list:
#             # 获取品牌
#             brand_name = li.xpath('./ul[@class="info-wrap"]/li[2]/span/em/text()').get()
#             print('品牌', brand_name)
#
#             # 获取店名
#             store_name = li.xpath('./ul[@class="info-wrap"]/li[1]/a/span/text()').get()
#             print('店名',store_name)
#
#
#             # 电话号
#             telephone_number = li.xpath('./ul[@class="info-wrap"]/li[3]/span[@class="tel"]/text()').get()
#             print('电话',telephone_number)
#
#             # 获取地址
#             address_of_4s = li.xpath('./ul[@class="info-wrap"]/li[4]/span[@class="info-addr"]/text()').get()
#             if '重庆' in address_of_4s:
#                 continue
#             if '广州' in address_of_4s:
#                 continue
#             print('地址', address_of_4s)
#             print('++++++++++++++++++++++++++++++++++++分割线++++++++++++++++++++++++++++++++++++')
#
#             # 获取城市
#             city_name = resp_data.xpath('//span[@id="change-city"]/span[@class="city-now"]/text()').get()
#             print('城市', city_name)
#
#             # Excel 存储
#             line = [brand_name, store_name, telephone_number,address_of_4s,city_name]
#             ws.append(line)
#             wb.save(r'C:\Users\liuyechun\Desktop\汽车之家-河北.xlsx')

















































